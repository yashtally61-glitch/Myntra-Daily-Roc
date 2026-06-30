import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Seller Reconciliation",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;600&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

section[data-testid="stSidebar"] {
    background: #0f172a !important;
    border-right: 1px solid #1e293b;
}
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] .stMarkdown li,
section[data-testid="stSidebar"] .stMarkdown span,
section[data-testid="stSidebar"] .element-container p,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stMultiSelect label,
section[data-testid="stSidebar"] .stDateInput label,
section[data-testid="stSidebar"] .stFileUploader label,
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3,
section[data-testid="stSidebar"] .upload-hint {
    color: #e2e8f0 !important;
}
section[data-testid="stSidebar"] .upload-hint b { color: #7dd3fc !important; }
section[data-testid="stSidebar"] [data-baseweb="tag"] { background: #1e40af !important; }
section[data-testid="stSidebar"] [data-baseweb="tag"] span { color: #ffffff !important; }

.main .block-container { padding-top: 1.5rem; max-width: 1400px; }

.metric-card {
    background: #ffffff; border: 1px solid #e2e8f0;
    border-radius: 12px; padding: 1.1rem 1.25rem;
    box-shadow: 0 2px 6px rgba(0,0,0,.07);
}
.metric-label { font-size:.72rem; font-weight:700; color:#475569; text-transform:uppercase; letter-spacing:.07em; margin-bottom:.3rem; }
.metric-value { font-size:1.5rem; font-weight:700; color:#0f172a; font-family:'JetBrains Mono',monospace; }
.metric-value.positive { color:#15803d; }
.metric-value.negative { color:#b91c1c; }
.metric-sub { font-size:.72rem; color:#64748b; margin-top:.2rem; font-weight:500; }

.section-header {
    display:flex; align-items:center; gap:.6rem;
    font-size:1rem; font-weight:700; color:#0f172a;
    border-bottom:2px solid #e2e8f0; padding-bottom:.5rem; margin-bottom:1rem;
}

.info-box { background:#eff6ff; border-left:3px solid #3b82f6; border-radius:0 8px 8px 0; padding:.8rem 1rem; font-size:.83rem; color:#1e3a5f; font-weight:500; }
.info-box b { color:#1d4ed8; }
.info-box code { background:#dbeafe; color:#1e40af; padding:.1rem .3rem; border-radius:3px; }
.info-box-green { background:#f0fdf4; border-left:3px solid #16a34a; border-radius:0 8px 8px 0; padding:.8rem 1rem; font-size:.83rem; color:#14532d; font-weight:500; }
.info-box-green b { color:#15803d; }
.info-box-green code { background:#dcfce7; color:#166534; padding:.1rem .3rem; border-radius:3px; }
.warn-box { background:#fffbeb; border-left:3px solid #f59e0b; border-radius:0 8px 8px 0; padding:.8rem 1rem; font-size:.83rem; color:#78350f; font-weight:500; }
.success-box { background:#f0fdf4; border-left:3px solid #22c55e; border-radius:0 8px 8px 0; padding:.8rem 1rem; font-size:.83rem; color:#14532d; font-weight:500; }

.upload-hint { background:#1e293b; border:1px solid #334155; border-radius:8px; padding:.85rem 1rem; font-size:.78rem; line-height:1.8; color:#e2e8f0; margin-top:.5rem; }
.file-pill { display:inline-block; background:#1e40af; color:#fff !important; border-radius:999px; padding:.2rem .7rem; font-size:.72rem; font-weight:600; margin:.15rem .1rem; }

h1 { color:#0f172a !important; font-weight:800 !important; }
h2 { color:#1e293b !important; font-weight:700 !important; }
[data-testid="stMetricLabel"] { color:#334155 !important; font-weight:600 !important; }
[data-testid="stMetricValue"] { color:#0f172a !important; font-weight:700 !important; }
.stTabs [data-baseweb="tab"] { color:#334155 !important; font-weight:600; font-size:1rem; }
.stTabs [data-baseweb="tab"][aria-selected="true"] { color:#1d4ed8 !important; }
</style>
""", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
#  SHARED CONSTANTS
# ═════════════════════════════════════════════════════════════════════════════
DEFAULT_MARKETING_TYPE  = "%"
DEFAULT_MARKETING_VALUE = 3.0
DEFAULT_RETURN_CHARGES  = 45.0
DEFAULT_ROYALTY_TYPE    = "%"
DEFAULT_ROYALTY_VALUE   = 1.0
GST_RATE                = 0.18

REQUIRED_RATE_COLS = [
    "Brand Name", "Category",
    "Lower Limit Commision", "Upper Limit Commision", "Commision Charge",
    "GT Lower Limit", "GT Upper Limit", "GT Charges",
    "Lower Limit Fixed Fee", "Upper Limit Fixed Fee", "Fix Fee",
]

_thin_blue  = Side(style="thin", color="B8CCE4")
_thin_green = Side(style="thin", color="BBF7D0")


# ═════════════════════════════════════════════════════════════════════════════
#  ███╗   ███╗██╗   ██╗███╗   ██╗████████╗██████╗  █████╗
#  ████╗ ████║╚██╗ ██╔╝████╗  ██║╚══██╔══╝██╔══██╗██╔══██╗
#  ██╔████╔██║ ╚████╔╝ ██╔██╗ ██║   ██║   ██████╔╝███████║
#  ██║╚██╔╝██║  ╚██╔╝  ██║╚██╗██║   ██║   ██╔══██╗██╔══██║
#  ██║ ╚═╝ ██║   ██║   ██║ ╚████║   ██║   ██║  ██║██║  ██║
#  ╚═╝     ╚═╝   ╚═╝   ╚═╝  ╚═══╝   ╚═╝   ╚═╝  ╚═╝╚═╝  ╚═╝  SECTION
# ═════════════════════════════════════════════════════════════════════════════

def compute_brand_charges(brand, SP, V, brand_cfg):
    cfg      = brand_cfg.get(brand, {})
    mkt_type = cfg.get("mkt_type", DEFAULT_MARKETING_TYPE)
    mkt_val  = cfg.get("mkt_val",  DEFAULT_MARKETING_VALUE)
    ret_ch   = cfg.get("return",   DEFAULT_RETURN_CHARGES)
    roy_type = cfg.get("roy_type", DEFAULT_ROYALTY_TYPE)
    roy_val  = cfg.get("roy_val",  cfg.get("roy_pct", DEFAULT_ROYALTY_VALUE))
    marketing = round(SP * (mkt_val / 100), 2) if mkt_type == "%" else round(float(mkt_val), 2)
    return_ch = round(float(ret_ch), 2)
    royalty   = round(V * (roy_val / 100), 2)  if roy_type == "%" else round(float(roy_val), 2)
    return marketing, return_ch, royalty


def validate_slab_ranges(rates):
    issues = []
    pairs = [
        ("Lower Limit Commision", "Upper Limit Commision", "Commision Charge", "Commission"),
        ("GT Lower Limit",        "GT Upper Limit",        "GT Charges",       "GT"),
        ("Lower Limit Fixed Fee", "Upper Limit Fixed Fee", "Fix Fee",          "Fixed Fee"),
    ]
    for (brand, cat), g in rates.groupby(["Brand Name", "Category"]):
        for lo_col, hi_col, val_col, label in pairs:
            gg = g.dropna(subset=[lo_col, hi_col]).sort_values(lo_col).reset_index()
            if gg.empty:
                continue
            for i in range(len(gg)):
                cur = gg.iloc[i]
                if cur[hi_col] < cur[lo_col]:
                    issues.append(dict(brand=brand, category=cat, kind="INVALID", column=label,
                        detail=f"row has {lo_col}={cur[lo_col]} > {hi_col}={cur[hi_col]}"))
                if i + 1 < len(gg):
                    nxt = gg.iloc[i + 1]
                    if nxt[lo_col] > cur[hi_col]:
                        issues.append(dict(brand=brand, category=cat, kind="GAP", column=label,
                            detail=f"no bracket covers {cur[hi_col]} – {nxt[lo_col]}"))
                    elif nxt[lo_col] < cur[hi_col]:
                        v1, v2 = cur[val_col], nxt[val_col]
                        if v1 != v2:
                            issues.append(dict(brand=brand, category=cat, kind="CONFLICT", column=label,
                                detail=(f"[{cur[lo_col]}-{cur[hi_col]}) and [{nxt[lo_col]}-{nxt[hi_col]}) "
                                        f"overlap with different values ({v1} vs {v2})")))
    return issues


# ── Myntra File Loaders ───────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def myntra_load_slab(file_bytes):
    import openpyxl
    wb   = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws   = wb["Rates"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ValueError("'Rates' sheet is empty.")
    header    = [c for c in rows[0] if c is not None and str(c).strip()]
    n_cols    = len(header)
    data_rows = []
    for r in rows[1:]:
        r = list(r)[:n_cols] + [None] * max(0, n_cols - len(r))
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        data_rows.append(r[:n_cols])
    rates = pd.DataFrame(data_rows, columns=header)
    rates.columns = [str(c).strip() for c in rates.columns]
    missing = [c for c in REQUIRED_RATE_COLS if c not in rates.columns]
    if missing:
        raise ValueError(f"'Rates' sheet missing columns: {', '.join(missing)}")
    rates["Brand Name"] = rates["Brand Name"].astype(str).str.strip()
    rates["Category"]   = rates["Category"].astype(str).str.strip()
    for c in [c for c in REQUIRED_RATE_COLS if c not in ("Brand Name","Category")]:
        rates[c] = pd.to_numeric(rates[c], errors="coerce")
    slab_num_cols = ["Lower Limit Commision","Upper Limit Commision",
                     "GT Lower Limit","GT Upper Limit","Lower Limit Fixed Fee","Upper Limit Fixed Fee"]
    rates = rates.dropna(subset=slab_num_cols, how="all").reset_index(drop=True)
    slab_issues = validate_slab_ranges(rates)
    oms_map, yrn_map = {}, {}
    if "Replace Sku" in wb.sheetnames:
        for row in wb["Replace Sku"].iter_rows(min_row=2, values_only=True):
            row = list(row) + [None] * max(0, 5 - len(row))
            if row[2] and row[4]: oms_map[str(row[2]).strip()] = str(row[4]).strip()
            if row[1] and row[4]: yrn_map[str(row[1]).strip()] = str(row[4]).strip()
    pwn_map = {}
    for sn in wb.sheetnames:
        if sn.strip().lower().startswith("price we need"):
            for row in wb[sn].iter_rows(min_row=2, values_only=True):
                row = list(row) + [None] * 3
                if row[1] and row[2] is not None:
                    pwn_map[str(row[1]).strip()] = row[2]
            break
    closed_map = {}
    if "Closed" in wb.sheetnames:
        for row in wb["Closed"].iter_rows(min_row=2, values_only=True):
            row = list(row) + [None] * 3
            if row[1] and row[2] is not None:
                closed_map[str(row[1]).strip()] = row[2]
    return rates, oms_map, yrn_map, pwn_map, closed_map, slab_issues


@st.cache_data(show_spinner=False)
def myntra_load_fix_rate(file_bytes):
    import openpyxl
    wb  = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws  = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
    df = pd.DataFrame(rows[1:], columns=header)
    required = ["Brand Name","Style Id","Start Date","End Date","GT Charge","Commission","Fixed fee"]
    missing  = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Fix Rate sheet missing columns: {', '.join(missing)}")
    df = df.dropna(how="all").reset_index(drop=True)
    df["Brand Name"] = df["Brand Name"].astype(str).str.strip()
    df["Style Id"]   = pd.to_numeric(df["Style Id"],   errors="coerce")
    df["GT Charge"]  = pd.to_numeric(df["GT Charge"],  errors="coerce").fillna(0)
    df["Commission"] = pd.to_numeric(df["Commission"], errors="coerce").fillna(0)
    df["Fixed fee"]  = pd.to_numeric(df["Fixed fee"],  errors="coerce").fillna(0)
    def _to_date(v):
        if isinstance(v, datetime): return v.date()
        if v is None or (isinstance(v, str) and not v.strip()): return None
        try: return pd.to_datetime(v).date()
        except: return None
    df["Start Date"] = df["Start Date"].apply(_to_date)
    df["End Date"]   = df["End Date"].apply(_to_date)
    return df[df["Style Id"].notna()].reset_index(drop=True)


@st.cache_data(show_spinner=False)
def myntra_load_csv(files_bytes, filenames):
    frames = []
    for fb, fn in zip(files_bytes, filenames):
        try:
            d = pd.read_csv(BytesIO(fb))
            d["_source_file"] = fn
            frames.append(d)
        except Exception as e:
            st.warning(f"Could not read '{fn}': {e}")
    if not frames:
        raise ValueError("No valid CSV files could be read.")
    return pd.concat(frames, ignore_index=True)


# ── Myntra Slab + Reconcile ───────────────────────────────────────────────────

def _lookup(subset, lo_col, hi_col, val):
    valid = subset[subset[lo_col].notna() & subset[hi_col].notna()]
    m = valid[(valid[lo_col] <= val) & (valid[hi_col] > val)]
    if m.empty:
        m = valid[(valid[lo_col] <= val) & (valid[hi_col] >= val)]
    return m.iloc[0] if not m.empty else None


def get_charges_from_slab(rates, brand, cat, SP):
    try:
        brand_str = str(brand).strip()
        cat_str   = str(cat).strip().lower()
        SP        = float(SP)
    except Exception:
        return None, None, None, None, None
    if not brand_str or brand_str.lower() in ("nan","none",""):
        return None, None, None, None, None
    if not cat_str or cat_str in ("nan","none",""):
        return None, None, None, None, None
    sub = rates[(rates["Brand Name"] == brand_str) & (rates["Category"].str.lower() == cat_str)]
    if sub.empty:
        return None, None, None, None, None
    gt_row = _lookup(sub, "GT Lower Limit", "GT Upper Limit", SP)
    if gt_row is None: return None, None, None, None, None
    GT = float(gt_row["GT Charges"])
    V  = round(SP - GT, 2)
    fee_row  = _lookup(sub, "Lower Limit Fixed Fee", "Upper Limit Fixed Fee", SP)
    if fee_row is None: return None, None, None, None, None
    comm_row = _lookup(sub, "Lower Limit Commision", "Upper Limit Commision", V)
    if comm_row is None: return None, None, None, None, None
    comm_rate = float(comm_row["Commision Charge"])
    comm_amt  = round(comm_rate * V, 2)
    fixed_fee = float(fee_row["Fix Fee"])
    return GT, V, comm_rate, comm_amt, fixed_fee


def get_fix_rate_row(fix_df, brand, style_id, order_date):
    if fix_df is None or fix_df.empty: return None
    try:   sid = int(float(style_id))
    except: return None
    sub = fix_df[(fix_df["Brand Name"] == brand) & (fix_df["Style Id"] == sid)]
    if sub.empty: return None
    if order_date is not None:
        dated = sub[sub["Start Date"].notna() & sub["End Date"].notna() &
                    (sub["Start Date"] <= order_date) & (sub["End Date"] >= order_date)]
        if not dated.empty: return dated.iloc[0]
        undated = sub[sub["Start Date"].isna() | sub["End Date"].isna()]
        if not undated.empty: return undated.iloc[0]
        return None
    return sub.iloc[0]


def myntra_get_pwn(oms_map, yrn_map, pwn_map, closed_map, seller_sku, myntra_sku=""):
    key     = str(seller_sku).strip()
    mkey    = str(myntra_sku).strip()
    oms_sku = oms_map.get(key, key)
    if oms_sku == key and mkey:
        oms_sku = yrn_map.get(mkey, oms_sku)
    if oms_sku in closed_map: return closed_map[oms_sku], oms_sku, "Closed"
    if oms_sku in pwn_map:    return pwn_map[oms_sku],    oms_sku, "PWN"
    return None, oms_sku, None


def myntra_reconcile(rates, df, oms_map, yrn_map, pwn_map, closed_map,
                     fix_df=None, report_date=None, brand_filter=None, brand_cfg=None):
    df = df.copy()
    df["brand"]        = df["brand"].astype(str).str.strip()
    df["article type"] = df["article type"].astype(str).str.strip()
    if brand_filter: df = df[df["brand"].isin(brand_filter)].copy()
    if brand_cfg is None: brand_cfg = {}
    records = []
    for _, row in df.iterrows():
        try:   SP = float(row["final amount"])
        except: SP = 0.0
        brand = str(row.get("brand","") or "").strip()
        cat   = str(row.get("article type","") or "").strip()
        sid   = row.get("style id", None)
        fix_row = get_fix_rate_row(fix_df, brand, sid, report_date)
        if fix_row is not None:
            GT = float(fix_row["GT Charge"]); comm_amt = float(fix_row["Commission"])
            fixed_fee = float(fix_row["Fixed fee"]); V = round(SP - GT, 2)
            total_ch  = round(comm_amt + GT + fixed_fee, 2)
            myntra_pay = round(SP - total_ch, 2)
            marketing, return_ch, royalty = compute_brand_charges(brand, SP, V, brand_cfg)
            rec = dict(_GT=GT, _V=V, _comm_rate=None, _comm_amt=comm_amt,
                       _fixed_fee=fixed_fee, _total_charges=total_ch, _gst=0.0,
                       _myntra_payable=myntra_pay, _marketing=marketing,
                       _return_ch=return_ch, _royalty=royalty,
                       _slab_ok=True, _rate_source="Fix Rate")
        else:
            GT, V, comm_rate, comm_amt, fixed_fee = get_charges_from_slab(rates, brand, cat, SP)
            if GT is None:
                rec = dict(_GT=None,_V=None,_comm_rate=None,_comm_amt=None,_fixed_fee=None,
                           _total_charges=None,_gst=None,_myntra_payable=None,
                           _marketing=None,_return_ch=None,_royalty=None,
                           _slab_ok=False,_rate_source="No Match")
            else:
                total_ch   = round(comm_amt + GT + fixed_fee, 2)
                gst        = round((total_ch - GT) * GST_RATE, 2)
                myntra_pay = round(SP - total_ch - gst, 2)
                marketing, return_ch, royalty = compute_brand_charges(brand, SP, V, brand_cfg)
                rec = dict(_GT=GT, _V=V, _comm_rate=comm_rate, _comm_amt=comm_amt,
                           _fixed_fee=fixed_fee, _total_charges=total_ch, _gst=gst,
                           _myntra_payable=myntra_pay, _marketing=marketing,
                           _return_ch=return_ch, _royalty=royalty,
                           _slab_ok=True, _rate_source="Slab")
        pwn_price, oms_sku, pwn_source = myntra_get_pwn(
            oms_map, yrn_map, pwn_map, closed_map,
            row.get("seller sku code",""), row.get("myntra sku code",""))
        rec["_pwn"] = pwn_price; rec["_oms_sku"] = oms_sku; rec["_pwn_source"] = pwn_source
        records.append(rec)
    enrich = pd.DataFrame(records, index=df.index)
    return pd.concat([df.reset_index(drop=True), enrich.reset_index(drop=True)], axis=1)


# ── Myntra Excel Export ───────────────────────────────────────────────────────

M_HEADERS = [
    "order release id","order line id","STYLE ID","Myntra SKU Code","SKU",
    "Original SKU","Article Type","MRP","PWN+10%",
    "Marketing Charges","Return Charges","Royalty","Total Amount",
    "Commission Amount","GT Charges","Fixed Fee",
    "Total Charges","GST Amount","Myntra Payble Amount",
    "Rebate","Diffrence","Selling Price","Selling Price -GT Price",
    "Date","Brand","Order Status","Rate Source",
]
M_COL_W = [18,15,12,24,24,22,14,8,10,16,14,10,13,17,11,10,13,11,18,8,11,13,18,13,12,12,12]

_M_HF = PatternFill("solid", fgColor="1F4E79")
_M_YF = PatternFill("solid", fgColor="FCE4D6")
_M_GF = PatternFill("solid", fgColor="E2EFDA")
_M_AF = PatternFill("solid", fgColor="FFF9E6")
_M_BF = PatternFill("solid", fgColor="DEEAF1")
_M_A1 = PatternFill("solid", fgColor="EBF3FA")
_M_A2 = PatternFill("solid", fgColor="FFFFFF")
_M_TB = Border(left=Side(style="thin",color="B8CCE4"), right=Side(style="thin",color="B8CCE4"),
               top=Side(style="thin",color="B8CCE4"),  bottom=Side(style="thin",color="B8CCE4"))


def _mcell(ws, r, c, val, fill, font=None, fmt=None, align="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill; cell.border = _M_TB
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if font: cell.font = font
    if fmt:  cell.number_format = fmt
    return cell


def myntra_build_excel(result_df, report_date_str="", brand_cfg=None):
    if brand_cfg is None: brand_cfg = {}
    wb         = Workbook()
    all_brands = sorted(result_df["brand"].unique())

    for brand in all_brands:
        bdf   = result_df[result_df["brand"] == brand].reset_index(drop=True)
        ws    = wb.create_sheet(title=brand[:31])
        hfont = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        nfont = Font(size=9, name="Calibri")
        ofont = Font(size=9, name="Calibri", italic=True, color="C55A11")
        gfont = Font(bold=True, size=9, name="Calibri", color="375623")
        ffont = Font(bold=True, size=9, name="Calibri", color="7B4F00")

        cfg      = brand_cfg.get(brand, {})
        mkt_type = cfg.get("mkt_type", DEFAULT_MARKETING_TYPE)
        mkt_val  = cfg.get("mkt_val",  DEFAULT_MARKETING_VALUE)
        roy_type = cfg.get("roy_type", DEFAULT_ROYALTY_TYPE)
        roy_val  = cfg.get("roy_val",  cfg.get("roy_pct", DEFAULT_ROYALTY_VALUE))
        ret_ch   = cfg.get("return",   DEFAULT_RETURN_CHARGES)

        dynamic_headers = list(M_HEADERS)
        dynamic_headers[9]  = f"Marketing ({mkt_val}{'%' if mkt_type=='%' else '₹ flat'})"
        dynamic_headers[10] = f"Return Charges (₹{ret_ch:.0f})"
        dynamic_headers[11] = f"Royalty ({roy_val}{'%' if roy_type=='%' else '₹ flat'})"

        for ci, (h, w) in enumerate(zip(dynamic_headers, M_COL_W), 1):
            _mcell(ws, 1, ci, h, _M_HF, hfont)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 36

        for ri, (_, row) in enumerate(bdf.iterrows(), 2):
            is_fix = row.get("_rate_source") == "Fix Rate"
            alt    = _M_AF if is_fix else (_M_A1 if ri % 2 == 0 else _M_A2)
            SP     = row.get("final amount", 0)
            V      = row.get("_V")
            gst_v  = row.get("_gst")
            rfont  = ffont if is_fix else nfont

            vals = [
                row.get("order release id",""), row.get("order line id",""),
                row.get("style id",""),         row.get("myntra sku code",""),
                row.get("seller sku code",""),  row.get("_oms_sku",""),
                row.get("article type",""),     row.get("total mrp",""),
                row.get("_pwn"),
                row.get("_marketing"),          row.get("_return_ch"),
                row.get("_royalty"),            None,
                row.get("_comm_amt"),           row.get("_GT"),
                row.get("_fixed_fee"),          row.get("_total_charges"),
                gst_v if gst_v else None,       row.get("_myntra_payable"),
                0,                              None,
                SP,                             V,
                report_date_str or "—",         row.get("brand",""),
                row.get("order status",""),     row.get("_rate_source",""),
            ]
            num_ci = {10,11,12,13,14,15,16,17,18,19,20,21,22,23}
            for ci, val in enumerate(vals, 1):
                if ci == 9:
                    _mcell(ws, ri, ci, val,
                           alt if val is not None else _M_YF,
                           rfont if val is not None else ofont,
                           fmt="#,##0.00" if val is not None else None)
                elif ci == 19:
                    _mcell(ws, ri, ci, val, _M_GF, gfont, fmt="#,##0.00" if val is not None else None)
                elif ci in (21, 13):
                    _mcell(ws, ri, ci, None, _M_BF if ci==21 else alt, rfont)
                elif ci == 27:
                    _mcell(ws, ri, ci, val, _M_AF if is_fix else _M_A2, ffont if is_fix else nfont)
                else:
                    fmt = "#,##0.00" if ci in num_ci else ("#,##0" if ci == 8 else None)
                    _mcell(ws, ri, ci, val, alt, rfont, fmt)

            ws.cell(row=ri, column=13).value         = f'=IF(I{ri}="","",K{ri}+J{ri}+I{ri}+L{ri})'
            ws.cell(row=ri, column=13).number_format = "#,##0.00"
            ws.cell(row=ri, column=21).value         = f'=IF(I{ri}="","",S{ri}-M{ri}+T{ri})'
            ws.cell(row=ri, column=21).number_format = "#,##0.00"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(M_HEADERS))}1"
        nr = len(bdf) + 3
        nc = ws.cell(row=nr, column=1,
            value="🟡 Amber = Fix Rate.  ⬜ White/Blue = Slab.  🟠 Orange PWN = fill manually.")
        nc.font = Font(bold=True, italic=True, size=9, color="7B4F00", name="Calibri")
        nc.fill = PatternFill("solid", fgColor="FFF9E6")
        ws.merge_cells(f"A{nr}:{get_column_letter(len(M_HEADERS))}{nr}")

    # Summary sheet
    ws_s = wb.create_sheet(title="Summary", index=0)
    sh   = ["Brand","Orders","Fix Rate Orders","Selling Price",
            "GT Charges","Commission","Fixed Fee","Total Charges","GST","Myntra Payable"]
    sw   = [14,8,14,18,14,14,12,14,10,18]
    hf2  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    for ci,(h,w) in enumerate(zip(sh,sw),1):
        _mcell(ws_s, 1, ci, h, _M_HF, hf2)
        ws_s.column_dimensions[get_column_letter(ci)].width = w
    ws_s.row_dimensions[1].height = 30

    grand = {k:0.0 for k in ["sp","gt","cm","ff","tc","gst","pay","fix_n","n"]}
    for ri, brand in enumerate(all_brands, 2):
        bdf   = result_df[result_df["brand"] == brand]
        ok    = bdf[bdf["_slab_ok"] == True]
        fix_n = int((bdf["_rate_source"] == "Fix Rate").sum())
        n=len(bdf); sp=round(bdf["final amount"].sum(),2)
        gt=round(ok["_GT"].sum(),2); cm=round(ok["_comm_amt"].sum(),2)
        ff=round(ok["_fixed_fee"].sum(),2); tc=round(ok["_total_charges"].sum(),2)
        gst=round(ok["_gst"].sum(),2); pay=round(ok["_myntra_payable"].sum(),2)
        alt = _M_A1 if ri%2==0 else _M_A2
        nf=Font(size=9,name="Calibri"); bf=Font(bold=True,size=9,name="Calibri")
        for ci,v in enumerate([brand,n,fix_n,sp,gt,cm,ff,tc,gst,pay],1):
            _mcell(ws_s, ri, ci, v, alt, bf if ci==1 else nf,
                   fmt="#,##0.00" if ci>3 else None)
        for k,v in zip(["n","sp","gt","cm","ff","tc","gst","pay","fix_n"],[n,sp,gt,cm,ff,tc,gst,pay,fix_n]):
            grand[k] += v

    gr = len(all_brands)+2
    gf2 = Font(bold=True,color="FFFFFF",size=10,name="Calibri")
    gv  = ["GRAND TOTAL",grand["n"],grand["fix_n"],
           round(grand["sp"],2),round(grand["gt"],2),round(grand["cm"],2),
           round(grand["ff"],2),round(grand["tc"],2),round(grand["gst"],2),round(grand["pay"],2)]
    for ci,v in enumerate(gv,1):
        _mcell(ws_s, gr, ci, v, _M_HF, gf2, fmt="#,##0.00" if ci>3 else None)

    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    buf = BytesIO(); wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
#   █████╗ ███╗   ███╗ █████╗ ███████╗ ██████╗ ███╗   ██╗
#  ██╔══██╗████╗ ████║██╔══██╗╚══███╔╝██╔═══██╗████╗  ██║
#  ███████║██╔████╔██║███████║  ███╔╝ ██║   ██║██╔██╗ ██║
#  ██╔══██║██║╚██╔╝██║██╔══██║ ███╔╝  ██║   ██║██║╚██╗██║
#  ██║  ██║██║ ╚═╝ ██║██║  ██║███████╗╚██████╔╝██║ ╚████║
#  ╚═╝  ╚═╝╚═╝     ╚═╝╚═╝  ╚═╝╚══════╝ ╚═════╝ ╚═╝  ╚═══╝  SECTION
# ═════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def amz_load_slab(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sku_map = {}
    if "Replace Sku" in wb.sheetnames:
        for row in wb["Replace Sku"].iter_rows(min_row=2, values_only=True):
            row = list(row) + [None]*3
            seller = str(row[0]).strip() if row[0] else None
            oms    = str(row[1]).strip() if row[1] else None
            if seller and oms and seller != "None":
                sku_map[seller] = oms
    pwn_map = {}
    for sn in wb.sheetnames:
        if "price we need" in sn.lower():
            for row in wb[sn].iter_rows(min_row=2, values_only=True):
                row = list(row) + [None]*3
                oms_child = str(row[1]).strip() if row[1] else None
                price     = row[2]
                if oms_child and oms_child != "None" and price is not None:
                    try: pwn_map[oms_child] = float(price)
                    except: pass
            break
    closed_map = {}
    if "Closed" in wb.sheetnames:
        for row in wb["Closed"].iter_rows(min_row=2, values_only=True):
            row = list(row) + [None]*3
            oms_child = str(row[1]).strip() if row[1] else None
            price     = row[2]
            if oms_child and oms_child != "None" and price is not None:
                try: closed_map[oms_child] = float(price)
                except: pass
    return sku_map, pwn_map, closed_map


@st.cache_data(show_spinner=False)
def amz_load_csv(file_bytes, filename):
    try:
        df = pd.read_csv(BytesIO(file_bytes), skiprows=13, encoding="utf-8-sig")
    except Exception:
        df = pd.read_csv(BytesIO(file_bytes), skiprows=13, encoding="latin-1")
    df.columns = [c.strip() for c in df.columns]
    df["_source_file"] = filename
    num_cols = ["product sales","shipping credits","gift wrap credits","promotional rebates",
                "Total sales tax liable(GST before adjusting TCS)",
                "TCS-CGST","TCS-SGST","TCS-IGST","TDS (Section 194-O)",
                "selling fees","fba fees","other transaction fees","other","total"]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c].astype(str).str.replace(",",""), errors="coerce").fillna(0.0)
    df["date/time"] = pd.to_datetime(df["date/time"], errors="coerce", dayfirst=True)
    if pd.api.types.is_datetime64tz_dtype(df["date/time"]):
        df["date/time"] = df["date/time"].dt.tz_localize(None)
    df["Sku"]       = df["Sku"].astype(str).str.strip()
    return df


def amz_reconcile(df, sku_map, pwn_map, closed_map):
    df = df.copy()
    df["OMS Sku"] = df["Sku"].apply(lambda x: sku_map.get(str(x).strip(), str(x).strip()))
    gst_col = "Total sales tax liable(GST before adjusting TCS)"
    df["Total Sales Amount"] = (
        df["product sales"]
        + df["shipping credits"]
        + df["gift wrap credits"]
        + df["promotional rebates"]
        + df[gst_col]
    ).round(2)
    def _comm_pct(row):
        tsa = row["Total Sales Amount"]
        fee = row.get("selling fees", 0)
        if tsa and abs(tsa) > 0.01: return round(fee / tsa * 100, 2)
        return None

    df["Comm%"] = df.apply(_comm_pct, axis=1)

    def _get_pwn(seller_sku, oms_sku):
        # 1) Try the Seller SKU exactly as it appears in the transaction file
        s_key = str(seller_sku).strip()
        if s_key in closed_map: return closed_map[s_key],          closed_map[s_key], "Closed"
        if s_key in pwn_map:    return pwn_map[s_key] + 50,         pwn_map[s_key],    "PWN"
        # 2) Fall back to the OMS SKU (via Replace Sku mapping)
        o_key = str(oms_sku).strip()
        if o_key != s_key:
            if o_key in closed_map: return closed_map[o_key],       closed_map[o_key], "Closed"
            if o_key in pwn_map:    return pwn_map[o_key] + 50,      pwn_map[o_key],    "PWN"
        # 3) Not found anywhere -- needs manual entry
        return None, None, None

    pwn_data       = df.apply(lambda r: _get_pwn(r["Sku"], r["OMS Sku"]), axis=1)
    df["PWN+RS50"]    = pwn_data.apply(lambda x: x[0])   # per-unit PWN+10% + ₹50
    df["PWN+10%"]     = pwn_data.apply(lambda x: x[1])   # per-unit PWN+10% price
    df["_pwn_source"] = pwn_data.apply(lambda x: x[2])

    # Quantity-adjusted target price -- PWN+RS50 must be multiplied by quantity,
    # otherwise multi-quantity order lines get compared against a single-unit price.
    qty_num = pd.to_numeric(df["quantity"], errors="coerce").fillna(1)
    qty_num = qty_num.where(qty_num > 0, 1)
    df["_qty_num"]          = qty_num
    df["PWN+RS50 (Total)"]  = (df["PWN+RS50"] * qty_num).round(2)

    total_col         = pd.to_numeric(df["total"], errors="coerce")
    df["Difference"]  = (total_col - df["PWN+RS50 (Total)"]).round(2)
    df["_pwn_matched"]= df["_pwn_source"].notna()
    return df


# ── Amazon Excel Export ───────────────────────────────────────────────────────

AMZ_EXPORT_COLS = [
    "date/time","type","order id","Sku","OMS Sku","description",
    "quantity","fulfillment","order city","order state",
    "product sales","shipping credits","gift wrap credits","promotional rebates",
    "Total sales tax liable(GST before adjusting TCS)",
    "Total Sales Amount","selling fees","fba fees","other transaction fees","other",
    "TCS-CGST","TCS-SGST","TCS-IGST","TDS (Section 194-O)",
    "total","Comm%","PWN+10%","PWN+RS50","PWN+RS50 (Total)","Difference",
    "Transaction Status","Transaction Release Date","_pwn_source",
]
AMZ_COL_HEADERS = [
    "Date/Time","Type","Order ID","Seller SKU","OMS SKU","Description",
    "Qty","Fulfillment","City","State",
    "Product Sales","Shipping Credits","Gift Wrap","Promo Rebates","GST",
    "Total Sales Amt","Selling Fees","FBA Fees","Other Txn Fees","Other",
    "TCS-CGST","TCS-SGST","TCS-IGST","TDS (194-O)",
    "Net Total","Comm %","PWN+10% (unit)","PWN+RS50 (unit)","PWN+RS50 (Total)","Difference",
    "Txn Status","Release Date","PWN Source",
]
AMZ_COL_WIDTHS = [
    20,12,22,22,22,40,6,10,14,14,
    13,13,10,13,10,14,12,10,14,10,
    10,10,10,12,12,9,14,14,16,12,12,20,12,
]

_A_HF = PatternFill("solid", fgColor="14532D")
_A_PF = PatternFill("solid", fgColor="DCFCE7")
_A_UF = PatternFill("solid", fgColor="FEF9C3")
_A_NF = PatternFill("solid", fgColor="FEE2E2")
_A_A1 = PatternFill("solid", fgColor="F0FDF4")
_A_A2 = PatternFill("solid", fgColor="FFFFFF")
_A_TB = Border(left=Side(style="thin",color="BBF7D0"), right=Side(style="thin",color="BBF7D0"),
               top=Side(style="thin",color="BBF7D0"),  bottom=Side(style="thin",color="BBF7D0"))
AMZ_NUM_CI = {11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,27,28,29,30}


def _acell(ws, r, c, val, fill, font=None, fmt=None, align="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill; cell.border = _A_TB
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(c == 6))
    if font: cell.font = font
    if fmt:  cell.number_format = fmt
    return cell


def amz_build_excel(result_df, report_label=""):
    wb   = Workbook()
    hf   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    nf   = Font(size=9, name="Calibri")
    bf   = Font(bold=True, size=9, name="Calibri")

    # ── Summary ───────────────────────────────────────────────────────────────
    ws_s = wb.active; ws_s.title = "Summary"
    sum_headers = ["Type","Total Rows","PWN Matched","Unmatched",
                   "Product Sales","GST","Total Sales Amt",
                   "Selling Fees","FBA Fees","Other Txn Fees",
                   "TCS","TDS","Net Total","PWN+RS50 Target","Difference (Net-Target)"]
    sum_widths  = [18,10,12,10,14,10,15,12,10,14,10,10,12,16,22]
    for ci,(h,w) in enumerate(zip(sum_headers,sum_widths),1):
        _acell(ws_s, 1, ci, h, _A_HF, hf)
        ws_s.column_dimensions[get_column_letter(ci)].width = w
    ws_s.row_dimensions[1].height = 30

    grand = {k:0.0 for k in ["ps","gst","tsa","sf","fba","otf","tcs","tds","net","pwn","diff","rows","matched","unmatched"]}
    gst_col = "Total sales tax liable(GST before adjusting TCS)"
    for ri, txn_type in enumerate(sorted(result_df["type"].dropna().unique()), 2):
        sub   = result_df[result_df["type"] == txn_type]
        tcs   = sub[["TCS-CGST","TCS-SGST","TCS-IGST"]].sum().sum()
        tds   = sub.get("TDS (Section 194-O)", pd.Series(0)).sum()
        mtch  = int(sub["_pwn_matched"].sum()) if "_pwn_matched" in sub else 0
        rv    = [txn_type, len(sub), mtch, len(sub)-mtch,
                 round(sub["product sales"].sum(),2),
                 round(sub[gst_col].sum(),2) if gst_col in sub else 0,
                 round(sub["Total Sales Amount"].sum(),2),
                 round(sub["selling fees"].sum(),2), round(sub["fba fees"].sum(),2),
                 round(sub["other transaction fees"].sum(),2),
                 round(tcs,2), round(tds,2),
                 round(pd.to_numeric(sub["total"],errors="coerce").sum(),2),
                 round(sub["PWN+RS50 (Total)"].sum(),2) if "PWN+RS50 (Total)" in sub else 0,
                 round(sub["Difference"].sum(),2) if "Difference" in sub else 0]
        alt = _A_A1 if ri%2==0 else _A_A2
        for ci,v in enumerate(rv,1):
            _acell(ws_s, ri, ci, v, alt, nf, fmt="#,##0.00" if ci>4 else None)
        for k,v in zip(["rows","matched","unmatched","ps","gst","tsa","sf","fba","otf","tcs","tds","net","pwn","diff"], rv[1:]):
            grand[k] += (v or 0)

    gr = result_df["type"].dropna().nunique() + 2
    gv = ["GRAND TOTAL",int(grand["rows"]),int(grand["matched"]),int(grand["unmatched"]),
          round(grand["ps"],2),round(grand["gst"],2),round(grand["tsa"],2),
          round(grand["sf"],2),round(grand["fba"],2),round(grand["otf"],2),
          round(grand["tcs"],2),round(grand["tds"],2),round(grand["net"],2),
          round(grand["pwn"],2),round(grand["diff"],2)]
    for ci,v in enumerate(gv,1):
        _acell(ws_s, gr, ci, v, _A_HF, Font(bold=True,color="FFFFFF",size=10,name="Calibri"),
               fmt="#,##0.00" if ci>4 else None)
    ws_s.freeze_panes = "A2"

    # ── Per-type sheets ───────────────────────────────────────────────────────
    order_types = ["Order","Refund","Fulfilment Fee Refund","Adjustment",
                   "Service Fee","FBA Inventory Fee","Reimbursements","SAFE-T Reimbursement"]
    for txn_type in order_types:
        sub = result_df[result_df["type"] == txn_type].copy()
        if sub.empty: continue
        ws  = wb.create_sheet(title=txn_type[:28].replace("/","-"))
        for ci,(h,w) in enumerate(zip(AMZ_COL_HEADERS,AMZ_COL_WIDTHS),1):
            _acell(ws, 1, ci, h, _A_HF, hf)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 28
        for ri,(_, row) in enumerate(sub.iterrows(),2):
            pwn_src = row.get("_pwn_source")
            base_fill = _A_NF if pwn_src=="Closed" else (_A_PF if pwn_src=="PWN" else (_A_UF if ri%2==0 else _A_A2))
            diff_val  = row.get("Difference")
            diff_font = Font(size=9,name="Calibri",bold=True,
                             color="15803D" if (diff_val is not None and not np.isnan(float(diff_val if diff_val is not None else 0)) and diff_val>=0) else "B91C1C")
            for ci, col in enumerate(AMZ_EXPORT_COLS, 1):
                val = row.get(col)
                fmt = "#,##0.00" if ci in AMZ_NUM_CI else None
                if isinstance(val, pd.Timestamp):
                    val = val.tz_localize(None) if val.tzinfo is not None else val
                    val = val.to_pydatetime() if not pd.isnull(val) else None
                    fmt = "dd-mmm-yyyy hh:mm"
                elif val is not None and isinstance(val, float) and np.isnan(val):
                    val = None
                font = diff_font if ci==30 else nf
                if ci==6: _acell(ws, ri, ci, val, base_fill, nf, align="left")
                else: _acell(ws, ri, ci, val, base_fill, font, fmt)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(AMZ_EXPORT_COLS))}1"
        nr = len(sub)+3
        nc = ws.cell(row=nr,column=1,value="🟢 Green=PWN matched · 🟡 Yellow=No match · 🔴 Red=Closed · Difference=Net Total−PWN+RS50")
        nc.font = Font(bold=True,italic=True,size=9,color="14532D",name="Calibri")
        nc.fill = PatternFill("solid",fgColor="DCFCE7")
        ws.merge_cells(f"A{nr}:{get_column_letter(len(AMZ_EXPORT_COLS))}{nr}")

    # ── All Transactions ──────────────────────────────────────────────────────
    ws_all = wb.create_sheet(title="All Transactions")
    for ci,(h,w) in enumerate(zip(AMZ_COL_HEADERS,AMZ_COL_WIDTHS),1):
        _acell(ws_all, 1, ci, h, _A_HF, hf)
        ws_all.column_dimensions[get_column_letter(ci)].width = w
    for ri,(_, row) in enumerate(result_df.iterrows(),2):
        alt = _A_A1 if ri%2==0 else _A_A2
        for ci, col in enumerate(AMZ_EXPORT_COLS,1):
            val = row.get(col)
            fmt = "#,##0.00" if ci in AMZ_NUM_CI else None
            if isinstance(val, pd.Timestamp):
                    val = val.tz_localize(None) if val.tzinfo is not None else val
                    val = val.to_pydatetime() if not pd.isnull(val) else None
                    fmt = "dd-mmm-yyyy hh:mm"
            elif val is not None and isinstance(val, float) and np.isnan(val):
                val = None
            _acell(ws_all, ri, ci, val, alt, nf, fmt)
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(AMZ_EXPORT_COLS))}1"

    buf = BytesIO(); wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI  —  top-level platform tabs
# ═════════════════════════════════════════════════════════════════════════════

st.markdown("## 🛍️ Seller Reconciliation")
st.markdown("<p style='color:#64748b;margin-top:-.5rem;margin-bottom:1rem;'>Choose a platform tab to get started</p>", unsafe_allow_html=True)

tab_myntra, tab_amazon = st.tabs(["  🧾 Myntra  ", "  📦 Amazon  "])


# ═════════════════════════════════════════════════════════════════════════════
#  MYNTRA TAB
# ═════════════════════════════════════════════════════════════════════════════
with tab_myntra:

    # ── Sidebar inputs (shown only when Myntra tab active) ───────────────────
    # We render sidebar inputs outside the tab because Streamlit sidebar is global.
    # We use session state to know which tab is "active" via a workaround:
    # just show all sidebar inputs and label them by platform.

    st.markdown("### Myntra Seller Reconciliation")

    # File uploaders inside the tab itself (works cleanly in Streamlit)
    mc1, mc2, mc3 = st.columns([1.5, 1.5, 1])
    with mc1:
        m_slab_file = st.file_uploader("① Slab.xlsx (Rates + Replace Sku)", type=["xlsx"], key="m_slab")
    with mc2:
        m_csv_files = st.file_uploader("② Orders CSV (one or more)", type=["csv"],
                                        key="m_csv", accept_multiple_files=True)
    with mc3:
        m_fix_file = st.file_uploader("③ Fix Rate Sheet (optional)", type=["xlsx"], key="m_fix")

    m_col_date, m_col_brand = st.columns([1, 2])
    with m_col_date:
        m_report_date = st.date_input("Report Date", value=datetime.today().date(), key="m_date")
    with m_col_brand:
        m_brand_filter = []
        if m_csv_files:
            try:
                _frames = []
                for f in m_csv_files:
                    _frames.append(pd.read_csv(f)); f.seek(0)
                _tmp = pd.concat(_frames, ignore_index=True)
                _tmp["brand"] = _tmp["brand"].astype(str).str.strip()
                _all_b = sorted(_tmp["brand"].unique().tolist())
                m_brand_filter = st.multiselect("Filter by Brand", options=_all_b,
                                                default=_all_b, key="m_brand_filter")
            except Exception:
                pass

    st.markdown("""<div class="upload-hint" style="margin-bottom:1rem;">
<b>Priority per order:</b> 1. Fix Rate (Brand + Style ID + Date) &nbsp;·&nbsp;
2. Slab (Brand + Category + SP range)<br>
<b>Fix Rate:</b> Payable = SP − GT − Comm − Fee (GST embedded) &nbsp;·&nbsp;
<b>Slab:</b> V=SP−GT · Comm=rate(V)×V · GST=(Comm+Fee)×18%
</div>""", unsafe_allow_html=True)

    if not m_slab_file or not m_csv_files:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("""<div class="info-box"><b>Step 1 – Slab.xlsx</b><br>
Rate table with a <code>Rates</code> sheet — Brand Name, Category, slab limits.</div>""",
                unsafe_allow_html=True)
        with c2:
            st.markdown("""<div class="info-box"><b>Step 2 – Orders CSV (one or more)</b><br>
Myntra Seller Orders export — <code>brand</code>, <code>article type</code>,
<code>final amount</code>, <code>style id</code>.</div>""", unsafe_allow_html=True)
    else:
        # ── Load Slab ─────────────────────────────────────────────────────────
        with st.spinner("Loading slab rates…"):
            try:
                m_rates, m_oms_map, m_yrn_map, m_pwn_map, m_closed_map, m_slab_issues = \
                    myntra_load_slab(m_slab_file.read())
            except Exception as e:
                st.error(f"❌ Slab error: {e}"); st.stop()

        if m_slab_issues:
            n_conflict = sum(1 for i in m_slab_issues if i["kind"] == "CONFLICT")
            n_gap      = sum(1 for i in m_slab_issues if i["kind"] in ("GAP","INVALID"))
            with st.expander(
                f"⚠️ Slab data issues — {len(m_slab_issues)} "
                f"({n_conflict} conflict{'s' if n_conflict!=1 else ''}, "
                f"{n_gap} gap/invalid) — click to review", expanded=True):
                st.markdown("""<div class="warn-box">
<b>CONFLICT</b> = overlapping brackets with different charges (wrong amount may be picked).
<b>GAP/INVALID</b> = price range with no matching bracket (shows as "Unmatched").
Fix the Rates sheet and re-upload.</div>""", unsafe_allow_html=True)
                st.dataframe(pd.DataFrame(m_slab_issues), use_container_width=True, hide_index=True)

        # ── Load Fix Rate ─────────────────────────────────────────────────────
        m_fix_df = None
        if m_fix_file:
            with st.spinner("Loading fix rates…"):
                try:
                    m_fix_df = myntra_load_fix_rate(m_fix_file.read())
                    n_combos = m_fix_df[["Brand Name","Style Id"]].drop_duplicates().shape[0]
                    st.success(f"✅ Fix Rate loaded — {len(m_fix_df)} rows · {n_combos} combos")
                except Exception as e:
                    st.warning(f"Fix Rate not loaded: {e}")

        # ── Load CSV ──────────────────────────────────────────────────────────
        with st.spinner("Reading orders…"):
            try:
                files_bytes = [f.read() for f in m_csv_files]
                filenames   = [f.name   for f in m_csv_files]
                m_orders_df = myntra_load_csv(files_bytes, filenames)
                m_orders_df["brand"]        = m_orders_df["brand"].astype(str).str.strip()
                m_orders_df["article type"] = m_orders_df["article type"].astype(str).str.strip()
                m_orders_df["final amount"] = pd.to_numeric(m_orders_df["final amount"], errors="coerce").fillna(0)
                m_orders_df["seller price"] = pd.to_numeric(m_orders_df["seller price"], errors="coerce").fillna(0)
                m_orders_df["style id"]     = pd.to_numeric(m_orders_df["style id"],     errors="coerce")
            except Exception as e:
                st.error(f"❌ CSV error: {e}"); st.stop()

        if len(m_csv_files) > 1:
            pills = " ".join(f'<span class="file-pill">{fn}</span>' for fn in filenames)
            st.markdown(f'<div class="info-box">📂 <b>{len(m_csv_files)} CSV files merged</b> '
                        f'— {len(m_orders_df):,} total rows<br>{pills}</div>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

        # ── Brand Charges Config ──────────────────────────────────────────────
        m_brands_in_data = sorted([b for b in m_orders_df["brand"].dropna().unique()
                                   if b and b.lower() not in ("nan","none","")])

        with st.expander("⚙️ Brand Charges Configuration", expanded=False):
            st.markdown("""<div class="info-box" style="margin-bottom:1rem;">
Set <b>Marketing</b>, <b>Return Charges</b> and <b>Royalty</b> per brand.
Pre-filled with defaults (Marketing 3%, Return ₹45, Royalty 1%).</div>""", unsafe_allow_html=True)

            m_brand_cfg = {}
            hcols = st.columns([1.8,1.4,1.3,1.5,1.4,1.3])
            for h,hc in zip(["Brand","Marketing Type","Marketing Value",
                              "Return Charges (₹)","Royalty Type","Royalty Value"], hcols):
                hc.markdown(f"**{h}**")
            st.markdown("<hr style='margin:.3rem 0 .6rem 0;border-color:#e2e8f0'>", unsafe_allow_html=True)

            for brand in m_brands_in_data:
                cols = st.columns([1.8,1.4,1.3,1.5,1.4,1.3])
                with cols[0]:
                    st.markdown(f"<div style='padding:.45rem 0;font-weight:600;color:#1e293b'>{brand}</div>", unsafe_allow_html=True)
                with cols[1]:
                    mkt_type = st.selectbox("mkt_type",["%","Flat ₹"], key=f"m_mkt_type_{brand}", label_visibility="collapsed")
                with cols[2]:
                    mkt_val  = st.number_input("mkt_val", min_value=0.0, value=DEFAULT_MARKETING_VALUE,
                                               step=0.5, format="%.2f", key=f"m_mkt_val_{brand}", label_visibility="collapsed")
                with cols[3]:
                    ret_ch   = st.number_input("return", min_value=0.0, value=DEFAULT_RETURN_CHARGES,
                                               step=1.0, format="%.2f", key=f"m_ret_{brand}", label_visibility="collapsed")
                with cols[4]:
                    roy_type = st.selectbox("roy_type",["%","Flat ₹"], key=f"m_roy_type_{brand}", label_visibility="collapsed")
                with cols[5]:
                    roy_val  = st.number_input("roy_val", min_value=0.0, value=DEFAULT_ROYALTY_VALUE,
                                               step=0.1, format="%.2f", key=f"m_roy_val_{brand}", label_visibility="collapsed")
                m_brand_cfg[brand] = {"mkt_type":mkt_type,"mkt_val":mkt_val,"return":ret_ch,"roy_type":roy_type,"roy_val":roy_val}

            preview_rows = [{"Brand":b,
                             "Marketing": f"{c['mkt_val']}% of SP" if c["mkt_type"]=="%"  else f"₹{c['mkt_val']:.2f} flat",
                             "Return (₹)":f"₹{c['return']:.2f}",
                             "Royalty":   f"{c['roy_val']}% of V"  if c["roy_type"]=="%" else f"₹{c['roy_val']:.2f} flat"}
                            for b,c in m_brand_cfg.items()]
            if preview_rows:
                st.markdown("**Current configuration:**")
                st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

        # ── Reconcile ─────────────────────────────────────────────────────────
        with st.spinner("Reconciling…"):
            m_result = myntra_reconcile(
                m_rates, m_orders_df, m_oms_map, m_yrn_map, m_pwn_map, m_closed_map,
                fix_df=m_fix_df, report_date=m_report_date,
                brand_filter=m_brand_filter if m_brand_filter else None,
                brand_cfg=m_brand_cfg)

        m_ok_df    = m_result[m_result["_slab_ok"] == True]
        m_bad_df   = m_result[m_result["_slab_ok"] == False]
        m_fix_rows = int((m_result["_rate_source"] == "Fix Rate").sum())

        # ── KPIs ──────────────────────────────────────────────────────────────
        total_orders  = len(m_result)
        total_sp      = m_result["final amount"].sum()
        total_payable = m_ok_df["_myntra_payable"].sum()
        total_tc_gst  = m_ok_df["_total_charges"].sum() + m_ok_df["_gst"].sum()
        unmatched     = len(m_bad_df)

        cols = st.columns(6)
        kpis = [
            ("Total Orders",        f"{total_orders:,}",       None, f"{len(m_csv_files)} file(s) merged"),
            ("Total Selling Price",  f"₹{total_sp:,.0f}",      None, "sum of final amount"),
            ("Myntra Payable",       f"₹{total_payable:,.0f}", None, "after all deductions"),
            ("Total Charges+GST",    f"₹{total_tc_gst:,.0f}", None, "platform fees"),
            ("Fix Rate Applied",     f"{m_fix_rows}/{total_orders}",
             "positive" if m_fix_rows>0 else None, "style-level overrides"),
            ("Unmatched Rows",       f"{unmatched}",
             "negative" if unmatched>0 else "positive", "no slab/fix rate"),
        ]
        for col, (label, value, cls, sub) in zip(cols, kpis):
            with col:
                st.markdown(f"""<div class="metric-card">
  <div class="metric-label">{label}</div>
  <div class="metric-value {cls or ''}">{value}</div>
  <div class="metric-sub">{sub}</div>
</div>""", unsafe_allow_html=True)

        if m_fix_rows > 0:
            st.markdown(f"""<br><div class="warn-box">
🔒 <b>{m_fix_rows} orders</b> used Fix Rate — GT/Commission/Fixed Fee from Fix Rate sheet.</div>""",
                unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Brand Summary ──────────────────────────────────────────────────────
        st.markdown('<div class="section-header">📊 Brand Summary</div>', unsafe_allow_html=True)
        summary_rows = []
        for brand in sorted(m_result["brand"].unique()):
            bdf  = m_result[m_result["brand"] == brand]
            bok  = bdf[bdf["_slab_ok"] == True]
            fx_n = int((bdf["_rate_source"] == "Fix Rate").sum())
            cfg  = m_brand_cfg.get(brand, {})
            mt   = cfg.get("mkt_type", DEFAULT_MARKETING_TYPE)
            mv   = cfg.get("mkt_val",  DEFAULT_MARKETING_VALUE)
            rt   = cfg.get("roy_type", DEFAULT_ROYALTY_TYPE)
            rv   = cfg.get("roy_val",  DEFAULT_ROYALTY_VALUE)
            summary_rows.append({
                "Brand":              brand,
                "Orders":             len(bdf),
                "Fix Rate Orders":    fx_n,
                "Marketing Rate":     f"{mv}% of SP" if mt=="%" else f"₹{mv:.2f} flat",
                "Return Charges (₹)": cfg.get("return", DEFAULT_RETURN_CHARGES),
                "Royalty":            f"{rv}% of V"  if rt=="%" else f"₹{rv:.2f} flat",
                "Selling Price (₹)":  round(bdf["final amount"].sum(),2),
                "GT Charges (₹)":     round(bok["_GT"].sum(),2),
                "Commission (₹)":     round(bok["_comm_amt"].sum(),2),
                "Total Charges (₹)":  round(bok["_total_charges"].sum(),2),
                "GST (₹)":            round(bok["_gst"].sum(),2),
                "Myntra Payable (₹)": round(bok["_myntra_payable"].sum(),2),
            })
        sum_df = pd.DataFrame(summary_rows)
        st.dataframe(sum_df.style
            .format({c:"{:,.2f}" for c in sum_df.columns if "₹" in c and c not in ("Return Charges (₹)","Royalty")}, na_rep="")
            .set_properties(**{"text-align":"center"}),
            use_container_width=True, hide_index=True)
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Order-level tabs ───────────────────────────────────────────────────
        brands_in_result = sorted(m_result["brand"].unique().tolist())
        if brands_in_result:
            st.markdown('<div class="section-header">📋 Order-Level Detail</div>', unsafe_allow_html=True)
            inner_tabs = st.tabs([f"  {b}  " for b in brands_in_result] + ["⚠️ Unmatched"])

            DISPLAY_COLS = {
                "order release id":"Order ID","order line id":"Line ID","style id":"Style ID",
                "seller sku code":"SKU","article type":"Category","order status":"Status",
                "_source_file":"Source File","_rate_source":"Rate Source",
                "final amount":"Selling Price ₹","_V":"SP−GT ₹","_comm_amt":"Commission ₹",
                "_GT":"GT Charges ₹","_fixed_fee":"Fixed Fee ₹","_total_charges":"Total Charges ₹",
                "_gst":"GST ₹","_myntra_payable":"Myntra Payable ₹",
                "_marketing":"Marketing ₹","_return_ch":"Return Charges ₹",
                "_royalty":"Royalty ₹","_pwn":"PWN+10% ₹",
            }
            NUM_KEYS = {"final amount","_V","_comm_amt","_GT","_fixed_fee",
                        "_total_charges","_gst","_myntra_payable","_marketing","_return_ch","_royalty","_pwn"}

            for inner_tab, brand in zip(inner_tabs[:-1], brands_in_result):
                with inner_tab:
                    bdf = m_result[m_result["brand"] == brand].copy()
                    ok  = bdf[bdf["_slab_ok"] == True]
                    mis = bdf[bdf["_slab_ok"] == False]
                    fx  = int((bdf["_rate_source"] == "Fix Rate").sum())
                    cfg = m_brand_cfg.get(brand, {})
                    mt  = cfg.get("mkt_type", DEFAULT_MARKETING_TYPE)
                    mv  = cfg.get("mkt_val",  DEFAULT_MARKETING_VALUE)
                    rc  = cfg.get("return",   DEFAULT_RETURN_CHARGES)
                    rt  = cfg.get("roy_type", DEFAULT_ROYALTY_TYPE)
                    rv  = cfg.get("roy_val",  DEFAULT_ROYALTY_VALUE)
                    st.markdown(f'<div class="info-box" style="margin-bottom:.75rem;">'
                                f'<b>Charges applied:</b> Marketing = {""+str(mv)+"%"+" of SP" if mt=="%"  else "₹"+str(mv)+" flat"}'
                                f' &nbsp;·&nbsp; Return = ₹{rc:.0f}'
                                f' &nbsp;·&nbsp; Royalty = {""+str(rv)+"%"+" of V" if rt=="%" else "₹"+str(rv)+" flat"}</div>',
                                unsafe_allow_html=True)
                    c1,c2,c3,c4,c5 = st.columns(5)
                    c1.metric("Orders",         len(bdf))
                    c2.metric("Selling Price",  f"₹{bdf['final amount'].sum():,.0f}")
                    c3.metric("Myntra Payable", f"₹{ok['_myntra_payable'].sum():,.0f}")
                    c4.metric("Fix Rate Rows",  fx)
                    c5.metric("Unmatched",      len(mis),
                              delta=str(len(mis)) if len(mis) else None, delta_color="inverse")

                    disp_cols  = {k:v for k,v in DISPLAY_COLS.items()
                                  if k != "_source_file" or len(m_csv_files) > 1}
                    display_df = bdf[[c for c in disp_cols if c in bdf.columns]].rename(columns=disp_cols)
                    num_dcols  = [disp_cols[k] for k in NUM_KEYS if k in disp_cols and k in bdf.columns]
                    for c in num_dcols:
                        if c in display_df.columns:
                            display_df[c] = pd.to_numeric(display_df[c], errors="coerce")

                    def _highlight(row):
                        if row.get("Rate Source") == "Fix Rate": return ["background-color:#fffbeb"]*len(row)
                        return [""]*len(row)

                    st.dataframe(display_df.style
                        .format({c:"{:,.2f}" for c in num_dcols if c in display_df.columns}, na_rep="")
                        .apply(_highlight, axis=1),
                        use_container_width=True, hide_index=True,
                        height=min(420, 45+35*len(display_df)))

                    if len(mis) > 0:
                        st.markdown(f'<div class="warn-box">⚠️ {len(mis)} rows had no matching slab or fix rate.</div>', unsafe_allow_html=True)
                        st.dataframe(mis[["seller sku code","style id","article type","final amount","order status"]],
                                     use_container_width=True, hide_index=True)

            with inner_tabs[-1]:
                if m_bad_df.empty:
                    st.markdown('<div class="success-box">✅ All rows matched.</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="warn-box">⚠️ {len(m_bad_df)} unmatched orders.</div>', unsafe_allow_html=True)
                    st.dataframe(m_bad_df[["brand","seller sku code","style id","article type","final amount","order status"]],
                                 use_container_width=True, hide_index=True)

        # ── Export ────────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown('<div class="section-header">⬇️ Export</div>', unsafe_allow_html=True)
        col_dl, col_info = st.columns([1,3])
        with col_dl:
            with st.spinner("Preparing Excel…"):
                m_xlsx = myntra_build_excel(m_result, report_date_str=str(m_report_date), brand_cfg=m_brand_cfg)
            st.download_button("📥 Download Myntra Reconciliation (.xlsx)",
                               data=m_xlsx, file_name="Myntra_Reconciliation.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with col_info:
            st.markdown("""<div class="info-box">One sheet per brand + <b>Summary</b> sheet.
Column headers show the <b>actual rate</b> applied for each brand.<br>
🟡 Amber = Fix Rate &nbsp;·&nbsp; ⬜ White/Blue = Slab &nbsp;·&nbsp; 🟠 Orange PWN = fill manually.</div>""",
                unsafe_allow_html=True)

        st.markdown("<br><br>", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
#  AMAZON TAB
# ═════════════════════════════════════════════════════════════════════════════
with tab_amazon:

    st.markdown("### Amazon Seller Reconciliation")

    ac1, ac2 = st.columns([1.5, 2])
    with ac1:
        a_slab_file = st.file_uploader(
            "① Amazon_Slab.xlsx (Replace Sku · Price We Need · Closed)",
            type=["xlsx"], key="a_slab")
    with ac2:
        a_csv_files = st.file_uploader(
            "② CustomUnifiedTransaction CSV (one or more)",
            type=["csv"], key="a_csv", accept_multiple_files=True)

    a_type_filter = st.multiselect(
        "Transaction types to show in detail tabs",
        options=["Order","Refund","Fulfilment Fee Refund","Adjustment",
                 "Service Fee","FBA Inventory Fee","Reimbursements","SAFE-T Reimbursement"],
        default=["Order","Refund","Fulfilment Fee Refund","Adjustment"],
        key="a_type_filter")

    st.markdown("""<div class="upload-hint" style="margin-bottom:1rem;">
<b>Logic:</b> Map Seller SKU → OMS SKU (Replace Sku) &nbsp;·&nbsp;
Total Sales Amt = Product Sales + GST &nbsp;·&nbsp;
PWN+RS50 = PWN+10% + ₹50 &nbsp;·&nbsp;
Difference = Net Total − (PWN+RS50 × Qty)
</div>""", unsafe_allow_html=True)

    if not a_slab_file or not a_csv_files:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("""<div class="info-box-green"><b>Step 1 – Amazon_Slab.xlsx</b><br>
Sheets: <code>Replace Sku</code> (SKU mapping) · <code>Price We Need Excel</code> (PWN+10%) · <code>Closed</code></div>""",
                unsafe_allow_html=True)
        with c2:
            st.markdown("""<div class="info-box-green"><b>Step 2 – Transaction CSV</b><br>
Amazon Custom Unified Transaction export from Seller Central. Multiple files merged automatically.</div>""",
                unsafe_allow_html=True)
    else:
        # ── Load Slab ─────────────────────────────────────────────────────────
        with st.spinner("Loading Amazon slab…"):
            try:
                a_sku_map, a_pwn_map, a_closed_map = amz_load_slab(a_slab_file.read())
            except Exception as e:
                st.error(f"❌ Slab error: {e}"); st.stop()

        st.success(f"✅ Slab loaded — {len(a_sku_map):,} SKU mappings · "
                   f"{len(a_pwn_map):,} PWN prices · {len(a_closed_map):,} Closed prices")

        # ── Load CSV(s) ───────────────────────────────────────────────────────
        with st.spinner("Reading transactions…"):
            a_frames = []
            for f in a_csv_files:
                try:
                    a_frames.append(amz_load_csv(f.read(), f.name))
                except Exception as e:
                    st.warning(f"Could not read '{f.name}': {e}")
            if not a_frames:
                st.error("❌ No valid CSV files loaded."); st.stop()
            a_raw_df = pd.concat(a_frames, ignore_index=True)

        st.markdown(f'<div class="info-box-green">📂 <b>{len(a_csv_files)} CSV file(s) merged</b> — '
                    f'{len(a_raw_df):,} total rows across {a_raw_df["type"].nunique()} transaction types</div>',
                    unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Reconcile ─────────────────────────────────────────────────────────
        with st.spinner("Reconciling…"):
            a_result = amz_reconcile(a_raw_df, a_sku_map, a_pwn_map, a_closed_map)

        # Keep only "Order" type transactions
        a_result = a_result[a_result["type"] == "Order"].reset_index(drop=True)

        # Drop orders where Total Sales Amount is 0
        a_result = a_result[a_result["Total Sales Amount"] != 0].reset_index(drop=True)

        a_orders_df = a_result[a_result["type"] == "Order"]
        gst_col     = "Total sales tax liable(GST before adjusting TCS)"

        # ── KPIs ──────────────────────────────────────────────────────────────
        a_total_orders    = len(a_orders_df)
        a_total_sales_amt = a_orders_df["Total Sales Amount"].sum()
        a_total_net       = pd.to_numeric(a_orders_df["total"], errors="coerce").sum()
        a_total_fba       = a_orders_df["fba fees"].sum()
        a_pwn_match       = int(a_orders_df["_pwn_matched"].sum())
        a_total_diff      = a_orders_df["Difference"].sum()

        cols = st.columns(6)
        a_kpis = [
            ("Total Orders",       f"{a_total_orders:,}",        None,       "type = Order"),
            ("Total Sales Amount", f"₹{a_total_sales_amt:,.0f}", None,       "product sales + GST"),
            ("Net Received",       f"₹{a_total_net:,.0f}",       None,       "after all Amazon deductions"),
            ("FBA Fees",           f"₹{a_total_fba:,.0f}",       "negative", "fulfilment charges"),
            ("PWN Matched",        f"{a_pwn_match}/{a_total_orders}",
             "positive" if a_pwn_match>0 else None, "orders with target price"),
            ("Total Difference",   f"₹{a_total_diff:,.0f}",
             "positive" if a_total_diff>=0 else "negative", "Net − (PWN+RS50 × Qty)"),
        ]
        for col, (label, value, cls, sub) in zip(cols, a_kpis):
            with col:
                st.markdown(f"""<div class="metric-card">
  <div class="metric-label">{label}</div>
  <div class="metric-value {cls or ''}">{value}</div>
  <div class="metric-sub">{sub}</div>
</div>""", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Transaction Type Summary ───────────────────────────────────────────
        st.markdown('<div class="section-header">📊 Transaction Type Summary</div>', unsafe_allow_html=True)

        def _color_diff(val):
            if pd.isna(val): return ""
            return "color:#15803d;font-weight:700" if val>=0 else "color:#b91c1c;font-weight:700"

        a_summary_rows = []
        for txn_type in sorted(a_result["type"].dropna().unique()):
            sub     = a_result[a_result["type"] == txn_type]
            gst_sum = sub[gst_col].sum() if gst_col in sub else 0
            tcs_sum = sub[["TCS-CGST","TCS-SGST","TCS-IGST"]].sum().sum()
            tds_sum = sub["TDS (Section 194-O)"].sum() if "TDS (Section 194-O)" in sub else 0
            mtch    = int(sub["_pwn_matched"].sum()) if "_pwn_matched" in sub else 0
            diff    = round(sub["Difference"].sum(),2) if "Difference" in sub else None
            a_summary_rows.append({
                "Type":               txn_type, "Rows": len(sub),
                "PWN Matched":        mtch,     "Unmatched": len(sub)-mtch,
                "Product Sales (₹)":  round(sub["product sales"].sum(),2),
                "GST (₹)":            round(gst_sum,2),
                "Total Sales Amt (₹)":round(sub["Total Sales Amount"].sum(),2),
                "Selling Fees (₹)":   round(sub["selling fees"].sum(),2),
                "FBA Fees (₹)":       round(sub["fba fees"].sum(),2),
                "TCS+TDS (₹)":        round(tcs_sum+tds_sum,2),
                "Net Total (₹)":      round(pd.to_numeric(sub["total"],errors="coerce").sum(),2),
                "Difference (₹)":     diff,
            })
        a_sum_df   = pd.DataFrame(a_summary_rows)
        a_money_c  = [c for c in a_sum_df.columns if "₹" in c]
        st.dataframe(a_sum_df.style
            .format({c:"{:,.2f}" for c in a_money_c}, na_rep="—")
            .map(_color_diff, subset=["Difference (₹)"])
            .set_properties(**{"text-align":"center"}),
            use_container_width=True, hide_index=True)
        st.markdown("<br>", unsafe_allow_html=True)

        # ── SKU-level summary (Orders) ─────────────────────────────────────────
        # ── SKU-level summary (Orders) ─────────────────────────────────────────
        st.markdown('<div class="section-header">🔍 SKU-Level Detail (Orders)</div>', unsafe_allow_html=True)
        if not a_orders_df.empty:
            sku_grp = a_orders_df.groupby(["Sku","OMS Sku","_pwn_source"]).agg(
                Qty=("quantity","sum"), Product_Sales=("product sales","sum"),
                Total_Sales_Amt=("Total Sales Amount","sum"),
                Selling_Fees=("selling fees","sum"), FBA_Fees=("fba fees","sum"),
                Net_Total=("total", lambda x: pd.to_numeric(x,errors="coerce").sum()),
                PWN10=("PWN+10%","first"), PWNRS50=("PWN+RS50","first"),
                PWNRS50_Total=("PWN+RS50 (Total)","sum"),
                Difference=("Difference","sum"),
            ).reset_index()
            sku_grp.columns = ["Seller SKU","OMS SKU","PWN Source","Qty",
                               "Product Sales ₹","Total Sales Amt ₹","Selling Fees ₹",
                               "FBA Fees ₹","Net Total ₹","PWN+10%","PWN+RS50 (unit)",
                               "PWN+RS50 (Total)","Difference ₹"]
            sku_grp = sku_grp.sort_values("Net Total ₹", ascending=False)
            money   = ["Product Sales ₹","Total Sales Amt ₹","Selling Fees ₹",
                       "FBA Fees ₹","Net Total ₹","PWN+10%","PWN+RS50 (unit)",
                       "PWN+RS50 (Total)","Difference ₹"]
            st.dataframe(sku_grp.style
                .format({c:"{:,.2f}" for c in money}, na_rep="—")
                .map(_color_diff, subset=["Difference ₹"])
                .set_properties(**{"text-align":"center"}),
                use_container_width=True, hide_index=True,
                height=min(500, 45+35*len(sku_grp)))

            unmatched_skus = sku_grp[sku_grp["PWN+10%"].isna()].reset_index(drop=True)

            if len(unmatched_skus) > 0:
                st.markdown(f'<div class="warn-box">⚠️ <b>{len(unmatched_skus)} SKU(s)</b> '
                            f'have no PWN price in the Slab file (checked Seller SKU then OMS SKU) — '
                            f'enter prices manually below.</div>',
                            unsafe_allow_html=True)

                if "a_manual_pwn" not in st.session_state:
                    st.session_state["a_manual_pwn"] = {}

                st.markdown("**Manual PWN+10% entry for unmatched SKUs:**")
                mc_h1, mc_h2, mc_h3, mc_h4 = st.columns([2.5, 2.5, 1.5, 2])
                mc_h1.markdown("**Seller SKU**")
                mc_h2.markdown("**OMS SKU**")
                mc_h3.markdown("**Net Total ₹**")
                mc_h4.markdown("**Manual PWN+10% ₹ (per unit)**")

                for _, urow in unmatched_skus.iterrows():
                    seller_sku = urow["Seller SKU"]
                    oms_sku    = urow["OMS SKU"]
                    map_key    = f"{seller_sku}||{oms_sku}"
                    mc1, mc2, mc3, mc4 = st.columns([2.5, 2.5, 1.5, 2])
                    mc1.write(seller_sku)
                    mc2.write(oms_sku)
                    mc3.write(f"₹{urow['Net Total ₹']:,.2f}")
                    val = mc4.number_input(
                        "manual_pwn", min_value=0.0,
                        value=float(st.session_state["a_manual_pwn"].get(map_key, 0.0)),
                        step=1.0, format="%.2f",
                        key=f"manual_pwn_{map_key}",
                        label_visibility="collapsed",
                    )
                    if val > 0:
                        st.session_state["a_manual_pwn"][map_key] = val
                    elif map_key in st.session_state["a_manual_pwn"]:
                        del st.session_state["a_manual_pwn"][map_key]

                manual_map = st.session_state["a_manual_pwn"]

                if manual_map:
                    def _apply_manual(row):
                        map_key = f"{row['Sku']}||{row['OMS Sku']}"
                        if pd.isna(row.get("PWN+10%")) and map_key in manual_map:
                            pwn10 = manual_map[map_key]
                            qty   = row.get("quantity", 1)
                            try: qty = float(qty)
                            except Exception: qty = 1.0
                            if not qty or qty <= 0: qty = 1.0
                            row["PWN+10%"]           = pwn10
                            row["PWN+RS50"]          = pwn10 + 50
                            row["PWN+RS50 (Total)"]  = round((pwn10 + 50) * qty, 2)
                            row["_pwn_source"]       = "Manual"
                            net = pd.to_numeric(row.get("total"), errors="coerce")
                            row["Difference"]        = round(net - row["PWN+RS50 (Total)"], 2) if pd.notna(net) else None
                            row["_pwn_matched"]      = True
                        return row

                    a_result    = a_result.apply(_apply_manual, axis=1)
                    a_orders_df = a_result[a_result["type"] == "Order"]
                    st.success(f"✅ Applied manual PWN+10% price to {len(manual_map)} SKU(s). "
                               "Tables and export below now reflect these changes.")
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Transaction detail tabs ────────────────────────────────────────────
        st.markdown('<div class="section-header">📋 Transaction Detail by Type</div>', unsafe_allow_html=True)
        active_types = [t for t in ["Order","Refund","Fulfilment Fee Refund","Adjustment",
                                     "Service Fee","FBA Inventory Fee","Reimbursements","SAFE-T Reimbursement"]
                        if t in a_result["type"].values]

        DISP_COLS = [
            ("date/time","Date"),("order id","Order ID"),("Sku","Seller SKU"),("OMS Sku","OMS SKU"),
            ("description","Description"),("quantity","Qty"),
            ("product sales","Product Sales ₹"),("Total Sales Amount","Total Sales Amt ₹"),
            ("selling fees","Selling Fees ₹"),("fba fees","FBA Fees ₹"),
            ("total","Net Total ₹"),("Comm%","Comm %"),
            ("PWN+10%","PWN+10% ₹"),("PWN+RS50","PWN+RS50 (unit) ₹"),
            ("PWN+RS50 (Total)","PWN+RS50 (Total) ₹"),
            ("Difference","Difference ₹"),("_pwn_source","PWN Source"),
            ("Transaction Status","Status"),
        ]
        NUM_DISP = {"Product Sales ₹","Total Sales Amt ₹","Selling Fees ₹","FBA Fees ₹",
                    "Net Total ₹","PWN+10% ₹","PWN+RS50 (unit) ₹","PWN+RS50 (Total) ₹",
                    "Difference ₹","Comm %"}

        if active_types:
            a_inner_tabs = st.tabs([f"  {t}  " for t in active_types])
            for a_tab, txn_type in zip(a_inner_tabs, active_types):
                with a_tab:
                    sub  = a_result[a_result["type"] == txn_type].copy()
                    mtch = int(sub["_pwn_matched"].sum()) if "_pwn_matched" in sub else 0
                    net  = pd.to_numeric(sub["total"], errors="coerce").sum()
                    diff = sub["Difference"].sum() if "Difference" in sub.columns else 0
                    c1,c2,c3,c4 = st.columns(4)
                    c1.metric("Rows",            len(sub))
                    c2.metric("Net Total",        f"₹{net:,.0f}")
                    c3.metric("PWN Matched",      f"{mtch}/{len(sub)}")
                    c4.metric("Total Difference", f"₹{diff:,.0f}")

                    avail = [(c, h) for c, h in DISP_COLS if c in sub.columns]
                    disp  = sub[[c for c,_ in avail]].rename(columns=dict(avail))
                    disp["Date"] = disp["Date"].apply(
                        lambda x: x.strftime("%d %b %Y") if isinstance(x, pd.Timestamp) else x)
                    for nc in NUM_DISP:
                        if nc in disp.columns:
                            disp[nc] = pd.to_numeric(disp[nc], errors="coerce")

                    disp = disp.loc[:, ~disp.columns.duplicated()].reset_index(drop=True)

                    col_cfg = {
                        c: st.column_config.NumberColumn(c, format="%.2f")
                        for c in NUM_DISP if c in disp.columns
                    }

                    st.dataframe(disp,
                        column_config=col_cfg,
                        use_container_width=True, hide_index=True,
                        height=min(450, 45+35*len(disp)))
        # ── Export ────────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown('<div class="section-header">⬇️ Export</div>', unsafe_allow_html=True)
        col_dl, col_info = st.columns([1,3])
        with col_dl:
            with st.spinner("Preparing Excel…"):
                a_xlsx = amz_build_excel(a_result, report_label=str(datetime.today().date()))
            st.download_button("📥 Download Amazon Reconciliation (.xlsx)",
                               data=a_xlsx,
                               file_name=f"Amazon_Reconciliation_{datetime.today().strftime('%Y%m%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with col_info:
            st.markdown("""<div class="info-box-green">
Excel export: <b>Summary</b> sheet + one sheet per transaction type + <b>All Transactions</b>.<br>
🟢 Green = PWN matched &nbsp;·&nbsp; 🟡 Yellow = no PWN &nbsp;·&nbsp; 🔴 Red = Closed price.<br>
<b>Difference</b> = Net Total − (PWN+RS50 × Qty).</div>""",
                unsafe_allow_html=True)
        st.markdown("<br><br>", unsafe_allow_html=True)
